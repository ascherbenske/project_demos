import pandas as pd
import numpy as np
import os
from sklearn.preprocessing import MinMaxScaler, LabelEncoder
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.ensemble import RandomForestClassifier
import xgboost as xgb
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from math import floor
from datetime import datetime

def main():    
    print('Running!')
    
    data_default = 'Uncombined'
    config_book_loc = f'{os.path.dirname(__file__)}\\excel_files\\App Config.xlsx'

    analysis = DataAnalysis()
    analysis.load_config_sheet(config_book_loc)

    if analysis.config_table['Training Data'].iloc[0] == data_default:
        df_training = analysis.convert_data_format(analysis.reqs_table_train,
                                                    analysis.crit_date_table_train,
                                                    'Related Critical Date ID',
                                                    'Critical Date ID')
        analysis._prep_data_hard_load(df_training, is_training=True)
    else:
        analysis._prep_data_easy_load(is_training=True)
    
    if analysis.config_table['Testing Data'].iloc[0] == data_default:
        df_testing = analysis.convert_data_format(analysis.reqs_table_eval,
                                                    analysis.crit_date_table_eval,
                                                    'Related Critical Date ID',
                                                    'Critical Date ID')
        analysis._prep_data_hard_load(df_testing, is_training=False)
        critical_date_ids = np.array(analysis.crit_date_table_eval['Critical Date ID'])
    else:
        analysis._prep_data_easy_load(is_training=False)
        critical_date_ids = np.array(analysis.raw_testing_data.cd_id)     
    
    X_train = analysis.training_data.drop(columns='delay').copy()
    y_train = analysis.training_data['delay'].copy().to_numpy()
    X_test = analysis.testing_data.copy()

    if 'delay' in list(X_test.columns):
        X_test = X_test.drop(columns='delay').copy()
    
    model = analysis.train_model(X_train, 
                                 y_train, 
                                 model_name=analysis.config_table['Model'].iloc[0])
    
    results_dict = analysis.predict_model(model, X_test)
    
    analysis.write_output(results_dict, 
                          critical_date_ids, 
                          analysis.config_table['Output Filename'].iloc[0])
    print('Complete!')

def export_df(df, file_name, dir_name=None):
    '''
    Export a dataframe to a CSV file

        Parameters:
            df (pd Dataframe): dataframe to export
            file_name (str): name of the file
            dir_name (str): directory name
    '''
    if dir_name is None:
        export_dir = os.path.dirname(__file__)
        export_dir = f'{export_dir}\\results'
    else:
        export_dir = f'{dir_name}\\'

    file_path = f'{export_dir}\\{file_name}.csv'
    df.to_csv(file_path, index=None)
    print(f'{file_name} is exported to {export_dir}')

class DataAnalysis(object):
    '''
    Overarching class that contains all of the relevant methods
    '''
    
    def __init__(self):
        self.training_data_filepath = None
        self.current_date = None
        self.raw_training_data = None
        self.raw_testing_data = None
        self.training_data = None
        self.testing_data = None
        self.ws_testing_data = None
        self.ws_training_data = None
        self.config_table = None
        self.crit_date_table_eval = None
        self.reqs_table_eval = None
        self.crit_date_table_train = None
        self.reqs_table_train = None
        self.model_to_use = None
        self.output_filename = None
        self.dir_name = os.path.dirname(__file__)
        self.status_list = ['not started', 
                            'in progress',
                            'not functional', 
                            'trial ready', 
                            'ops ready']
        self.xgb_dict = {}
        self.xgb_dict_reverse = {}

    def easy_load_data(self, data_filepath, is_training=True):
        '''
        Imports testing or training data without transforming it from [client's name] format

            Parameters:
                data_filepath (str): filepath of the data
                is_training (bool): True for if it's training data, false if testing

        '''
        file_path = f'{self.dir_name}\\{data_filepath}'
        
        if is_training:
            self.raw_training_data = pd.read_csv(file_path)
        else:
            self.raw_testing_data = pd.read_csv(file_path)
    
    def load_config_sheet(self, file_path):
        '''
        Loads the configuration excel workbook

            Parameters:
                file_path (str): filepath of the workbook
        '''
        
        wb = load_workbook(file_path, data_only=True)
        table_dict = {'tblConfig': 'Configuration',
                      'tblRequirements_Eval': 'Requirements_Eval',
                      'tblCriticalDates_Eval': 'CriticalDates_Eval',
                      'tblRequirements_Train': 'Requirements_Train',
                      'tblCriticalDates_Train': 'CriticalDates_Train',
                      'tblTesting': 'TestData',
                      'tblTraining': 'TrainingData'}
        
        for table, ws in table_dict.items():
            worksheet = wb[ws]
            temp_table = worksheet.tables[table]

            min_col, min_row, max_col, max_row = range_boundaries(temp_table.ref)

            data = []
            for row in worksheet.iter_rows(min_row=min_row, max_row=max_row,
                                    min_col=min_col, max_col=max_col,
                                    values_only=True):
                data.append(row)

            df = pd.DataFrame(data=data[1:], columns=data[0])
            self._load_config_tables(df, table)

    def _load_config_tables(self, data, table_name):        
        '''
        Assigns the tables from the configuration sheet to the corresponding attribute

            Parameters:
                data (pd DataFrame): dataframe to assign to an attribute
                table_name (str): name of the table in the workbook

        '''
        
        if table_name == 'tblConfig':
            self.config_table = data
        elif table_name == 'tblRequirements_Eval':
            self.reqs_table_eval = data
        elif table_name == 'tblCriticalDates_Eval':
            self.crit_date_table_eval = data
        elif table_name == 'tblRequirements_Train': 
            self.reqs_table_train = data
        elif table_name == 'tblCriticalDates_Train':
            self.crit_date_table_train = data
        elif table_name == 'tblTesting':
            self.raw_testing_data = data
        elif table_name == 'tblTraining':
            self.raw_training_data = data
    
    def convert_data_format(self, reqs_table, cd_table, as_weeks=True):
        '''
        Converts the requirements and the critical dates table to a format
        the app will read 

            Parameters:
                reqs_table (pd DataFrame): dataframe containing the requirements
                cd_table (pd DataFrame): dataframe containing the critical dates
                as_weeks (bool): whether the time period are weeks or not

            Returns:
                df (pd DataFrame): dataframe with converted values
        '''
       
        left_table = reqs_table
        right_table = cd_table
        left_id = 'Related Critical Date ID'
        right_id = 'Critical Date ID'
        
        df = pd.merge(left_table,
                      right_table,
                      how='outer',
                      left_on=left_id,
                      right_on=right_id)
        
        df = df.drop(columns=['Related Critical Date ID', 'Requirement'])
        df['Plan Date'] = pd.to_datetime(df['Plan Date'])
        df['time_until_cd'] = (df['Plan Date'] - df['Date Snapshot']).dt.days 
        
        if as_weeks:
            df['time_until_cd'] = np.floor(df['time_until_cd'] / 7)
        
        df_statuses = self._calculate_status_columns(df)
        
        df = pd.merge(left=df,
                      right=df_statuses,
                      on=['Critical Date ID', 'Date Snapshot'])
        
        df = df.drop(columns=['Req ID',
                              'Status',
                              'Readiness',
                              'Topic'])
        
        df = df.drop_duplicates().reset_index()
        df = self._get_previous_snapshot(df)

        return df        

    def _get_previous_snapshot(self, df):
        '''
        Gets the values from the previous snapshot, filters out first snapshot per CD

            Parameters:
                df (pd DataFrame): dataframe containing the periodic snapshots

            Returns:
                df (pd DataFrame): dataframe from previous snapshot
        '''

        fields_list = ['Date Snapshot', 
                       'not started_pct',
                       'in progress_pct',
                       'not functional_pct',
                       'trial ready_pct',
                       'ops ready_pct']
        
        prev_cols_list = [f'{x}_prev' for x in fields_list]

        shifted_df = df.copy()
        shifted_df[prev_cols_list] = df.groupby(by='Critical Date ID')[fields_list].shift(1)
        shifted_df = shifted_df.reset_index()
        shifted_df = shifted_df[shifted_df[prev_cols_list].notnull().all(1)]
        shifted_df = shifted_df.drop(columns=['level_0', 'index'])
        
        return shifted_df

    def _calculate_status_columns(self, df):
        '''
        Calculates the percentage of each status for a given CD and snapshot date

            Parameters:
                df (pd DataFrame): dataframe containing the periodic snapshots

            Returns:
                df (pd DataFrame): dataframe with new columns
        '''
        
        
        df_1 = df.groupby(by=['Date Snapshot', 'Critical Date ID']).count().reset_index()
        df_1 = df_1[['Date Snapshot', 
                     'Critical Date ID', 
                     'Status']].rename(columns={'Status': 'Full Count'})
        
        df_2 = df.groupby(by=['Date Snapshot', 
                              'Critical Date ID', 
                              'Status']).count().reset_index()
        df_2 = df_2[['Date Snapshot', 
                     'Critical Date ID', 
                     'Status',
                     'Readiness']].rename(columns={'Readiness': 'Status Count'})
        
        df_3 = pd.merge(left=df_1,
                        right=df_2,
                        on=['Critical Date ID', 'Date Snapshot'])
        
        df_3['Status Pct'] = np.round(df_3['Status Count'] / df_3['Full Count'], 4)
        df_3['Status'] = df_3['Status'].str.lower()

        df_status_pct = df_3[['Date Snapshot', 'Critical Date ID']].drop_duplicates()
        
        for status in self.status_list:
            df_temp = df_3.query('Status == @status').copy()
            new_field = f'{status}_pct'

            if df_temp.empty:
                df_status_pct[new_field] = 0.0
            else:
                df_status_pct = pd.merge(left=df_status_pct, 
                                right=df_temp,
                                on=['Date Snapshot', 'Critical Date ID'],
                                how='outer')
                df_status_pct = df_status_pct.rename(columns={'Status Pct': new_field})
                df_status_pct = df_status_pct.drop(columns=['Status',
                                                            'Full Count',
                                                            'Status Count'])
        
        return df_status_pct.fillna(0.0)
    
    def _prep_data_hard_load(self, data, is_training=True):
        '''
        Runs methods to clean and prep the data when loading it in [client's name] format

            Parameters:
                data (pd DataFrame): dataframe containing the data
                is_training (bool): whether it's training or testing data
        '''
        
        drop_col_list = ['Critical Date ID', 'Plan Date', 'Date Snapshot_prev', 'Date Snapshot']
        scale_col_list = ['n_issues', 'n_risks', 'time_until_cd']
        
        if is_training:
            self._clean_data(data, drop_cols=drop_col_list)
            self._rename_converted_cols(self.training_data)
            self._scale_data(self.training_data, scale_cols=scale_col_list)
        else:
            self._clean_data(data, drop_cols=drop_col_list, is_training_data=is_training)
            self._rename_converted_cols(self.testing_data, is_training_data=is_training)
            self._scale_data(self.training_data, scale_cols=scale_col_list, is_training_data=False)

    def _rename_converted_cols(self, data, is_training_data=True):
        '''
        Renames the columns of the prepped dataframe

            Parameters:
                data (pd DataFrame): dataframe containing the data
                is_training (bool): whether it's training or testing dat

            Returns:
                df (pd DataFrame): dataframe with new column names
        '''
        
        rename_dict = {'Num Issues': 'n_issues',
                       'Num Risks': 'n_risks',
                       'not started_pct': 'not_started',
                       'in progress_pct': 'in_progress',
                       'not functional_pct': 'not_functional',
                       'trial ready_pct': 'trial_ready',
                       'ops ready_pct': 'ops_ready',
                       'not started_pct_prev': 'not_started_prev',
                       'in progress_pct_prev': 'in_progress_prev',
                       'not functional_pct_prev': 'not_functional_prev',
                       'trial ready_pct_prev': 'trial_ready_prev',
                       'ops ready_pct': 'ops_ready_prev',
                       'Delay': 'delay'}
        
        df = data.rename(columns=rename_dict).copy()
        self._check_training_testing(df, is_training_data)

    def _prep_data_easy_load(self, is_training=True):
        '''
        Runs methods to clean and prep the data when loading it NOT in [client's name] format

            Parameters:
                data (pd DataFrame): dataframe containing the data
                is_training (bool): whether it's training or testing data
        '''
        
        if is_training:
            self._clean_data(self.raw_training_data)
            self._transform_data(self.training_data)
            self._scale_data(self.training_data)
        
        else:
            self._clean_data(self.raw_testing_data, is_training_data=is_training)
            self._transform_data(self.testing_data, is_training_data=is_training)
            self._scale_data(self.testing_data, is_training_data=is_training)

    def _scale_data(self, data, is_training_data=True, scale_cols=None):
        '''
        Scales selected columns

            Parameters:
                data (pd DataFrame): dataframe containing the data
                is_training (bool): whether it's training or testing data
                scale_cols (list of str): optional param to designate columns to scale
        '''
        
        df = data.copy()

        if scale_cols is None:
            scale_list = ['week_number',
                      'duration',
                      'start_week',
                      'weeks_until_req_end',
                      'weeks_until_proj_end']
        else:
            scale_list = scale_cols

        for feature in scale_list:
            df[feature] = (MinMaxScaler()
                                   .fit_transform(
                                       np.reshape(df[feature], (-1,1))))

        self._check_training_testing(df, is_training_data)

    def _transform_data(self, data, is_training_data=True):
        '''
        Creates new columns for timekeeping

            Parameters:
                data (pd DataFrame): dataframe containing the data
                is_training (bool): whether it's training or testing data
        '''
        
        df = data.copy()
        df['weeks_until_req_end'] = df['duration'] + df['start_week'] - df['week_number']
        df['weeks_until_proj_end'] = df['project_duration'] - df['week_number']

        self._check_training_testing(df, is_training_data)

    def _clean_data(self, data, is_training_data=True, drop_cols=None):
        '''
        Drops a set of columns

            Parameters:
                data (pd DataFrame): dataframe containing the data
                is_training (bool): whether it's training or testing data
                drop_cols (list of str): optional param to designate columns to drop
        '''
        
        if drop_cols == None:
            drop_list = ['ID', 
                     'weight_in_progress', 
                     'weight_not_functional', 
                     'weight_trial_ready',
                     'cd_id',
                     'project']
        else:
            drop_list = drop_cols

        col_list = list(data.columns)
        drop_list = [x for x in drop_list if x in col_list]
        
        data = data.drop(columns=drop_list)
        self._check_training_testing(data, is_training_data)
                
    def _check_training_testing(self, data, is_training_data=True):
        '''
        Wrapper method to assign a dataframe to either training data or testing

            Parameters:
                data (pd DataFrame): dataframe containing the data
                is_training (bool): whether it's training or testing data

            Returns:
                Assigns dataframe to either training_data or testing_data attributes
        '''
        
        if is_training_data:
            self.training_data = data
        else:
            self.testing_data = data

    def train_model(self, X_train, y_train, model_name='random forest', params=None):
        '''
        Wrapper method to train a model; will use hardcoded parameters if not passed in

            Parameters:
                X_train (pd DataFrame): dataframe containing the predictor data
                y_train (Iterable): response data
                model_name (str): name of the model to use
                params (dict): parameters to use for models

            Returns:
                A model trained using the given parameters
        '''
        
        model_lower = model_name.lower()
        
        if params is not None:
            param_dict = params
        elif model_lower == 'xgboost':
            param_dict = {'n_estimators': 200,
                          'max_depth': 65,
                          'grow_policy': 'lossguide',
                          'learning_rate': 0.09}
        elif model_lower == 'random forest':
            param_dict = {'n_estimators': 500,
                          'min_samples_leaf': 5,
                          'min_samples_split': 15,
                          'min_impurity_decrease': 0.05,
                          'max_depth': 85,
                          'criterion': 'entropy'}       
        elif model_lower == 'gradient boosting':
            param_dict = {'n_estimators': 400,
                          'min_samples_leaf': 5,
                          'min_samples_split': 5,
                          'min_impurity_decrease': 0.0,
                          'learning_rate': 0.05,
                          'max_depth': 5}
        elif model_lower == 'logistic regression':
            param_dict = {'penalty': 'elasticnet',
                          'solver': 'saga',
                          'l1_ratio': 0.5,
                          'C': 1.0}
            
        if model_lower == 'random forest':
            return RandomForestClassifier(**param_dict).fit(X_train, y_train)
        elif model_lower == 'gradient boosting':
            return GradientBoostingClassifier(**param_dict).fit(X_train, y_train)
        elif model_lower == 'xgboost':
            y_vals = LabelEncoder().fit_transform(y_train)
            return xgb.XGBClassifier(**param_dict).fit(X_train, y_vals)
        
    def predict_model(self, model, X_test):
        '''
        Gets the various required responses from the model and predicts it using X_test

            Parameters:
                model (scikit-learn model): trained model
                X_test (pd Dataframe): testing data to evaluate

            Returns:
                Dictionary with matrices of predictions, probabilities, and classes
        '''
        
        prediction = model.predict(X_test)
        probabilities = model.predict_proba(X_test)
        classes = model.classes_

        if self.model_to_use == 'xgboost':
            classes = np.vectorize(self.xgb_dict_reverse.get)(classes)        

        return {'y_predicted': prediction,
                'probabilities': probabilities,
                'y_possible': classes}
    
    def write_output(self, results_dict, cd_ids, file_name):
        '''
        Creates the formatted output file of the predicted classes and probabilities
        and exports it as a CSV

            Parameters:
                results_dict (dict): dict from predict_model method
                cd_ids (list of str): list of the Critical Date IDs
                file_name (str): filename to export
        '''
        
        dict_db = {'Critical Date ID': [],
                   'Predicted Delay': [],
                   'Prediction Probability': [],
                   'On Time Probability': [],
                   'Earlier Probability': [],
                   'Later Probability': []}
        
        on_time_idx = (results_dict['y_possible'] == 0).nonzero()[0][0]
        
        for i, y_pred in enumerate(list(results_dict['y_predicted'])):
            dict_db['Critical Date ID'].append(cd_ids[i])
            dict_db['Predicted Delay'].append(y_pred)
            
            class_idx = (results_dict['y_possible'] == y_pred).nonzero()[0][0]
            proba_row = results_dict['probabilities'][i]
            dict_db['Prediction Probability'].append(round(proba_row[class_idx], 2))
            dict_db['On Time Probability'].append(round(proba_row[on_time_idx], 2))

            if y_pred > 0:
                dict_db['Earlier Probability'].append(round(np.sum(proba_row[:class_idx]), 2))
            else:
                dict_db['Earlier Probability'].append(-1)
            
            if y_pred < np.max(results_dict['y_possible']):
                temp_idx = class_idx + 1
                dict_db['Later Probability'].append(round(np.sum(proba_row[temp_idx:]), 2))
            else:
                dict_db['Later Probability'].append(-1)  

        dir_name = os.path.dirname(__file__)
        dir_name = f'{dir_name}\\results'

        current_date = datetime.today().strftime('%Y-%m-%d')
        temp_file_name = f'{file_name}_{current_date}'

        export_df(pd.DataFrame(dict_db), temp_file_name, dir_name)

    def _days_to_weeks(self, date_1, date_2):
        '''
        Converts the absolute difference between two dates to weeks

            Parameters:
                date_1 (pd DateTime): first date
                date_2 (pd DateTime): second date
        '''
        
        days = abs((date_1 - date_2).days)
        return floor(days / 7)

if __name__ == '__main__':
    main()